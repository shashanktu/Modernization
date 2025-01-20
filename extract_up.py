# import re
# import os
# import openpyxl

# def extract_file_names(directory, excel_path):
#     # Define a regex pattern to match any file extension in 'src' or 'implementation' attributes
#     # file_pattern = re.compile(r'src=".*?([^/\\]+)\.([a-zA-Z0-9]+)"', re.IGNORECASE)
#     file_pattern = re.compile(r'[\w\-]+\.[a-zA-Z]{2,3}', re.IGNORECASE)
#     aqs_pattern = re.compile(r'<aqs:(\w+)', re.IGNORECASE)

#     # Dictionary to store file names grouped by extension
#     file_by_extension = {}
#     aqs_tags = set()

#     # Traverse the directory and process each .asp file
#     for root, _, files in os.walk(directory):
#         for file in files:
#             if file.lower().endswith('.asp'):
#                 file_path = os.path.join(root, file)
#                 with open(file_path, 'r') as f:
#                     content = f.read()

#                 for filename, extension in file_pattern.findall(content):
#                     if extension not in file_by_extension:
#                         file_by_extension[extension] = set()
#                     file_by_extension[extension].add(f"{filename}.{extension}")
                
#                 # Extract aqs tags
#                 aqs_tags.update(aqs_pattern.findall(content))

#     # Create an Excel workbook and add a worksheet
#     workbook = openpyxl.Workbook()
#     sheet = workbook.active
#     sheet.title = "Extracted Names"

#     # Write headers for the extracted data (file extensions)
#     headers = list(file_by_extension.keys()) + ["AQS Tags"]
#     sheet.append(headers)

#     # Determine the maximum number of rows needed for the file extensions and AQS tags
#     max_rows = max(len(files) for files in file_by_extension.values())  # Largest number of file names in a column
#     max_rows = max(max_rows, len(aqs_tags))  # Ensure there are enough rows for AQS tags

#     # Write file names and AQS tags into the Excel sheet
#     for i in range(max_rows):
#         row = []
        
#         # Add file names for each extension column
#         for extension in headers[:-1]:  # Exclude the "AQS Tags" column
#             file_list = list(file_by_extension.get(extension, []))
#             row.append(file_list[i] if i < len(file_list) else '')  # Add file name or empty cell
        
#         # Add AQS tags to the last column
#         row.append(list(aqs_tags)[i] if i < len(aqs_tags) else '')  # Add AQS tag or empty cell
#         sheet.append(row)

#     # Save the Excel workbook
#     workbook.save(excel_path)

# # Replace 'your_directory' with the path to the directory containing your .asp files
# # Replace 'output.xlsx' with the desired path for the Excel file
# extract_file_names(r'C:\Users\ShashankTudum\Downloads\AQS_VBScript\AQS_VBScript', r'C:\Users\ShashankTudum\Downloads\shashank-output\output.xlsx')


""" Third way"""
# import os
# import re
# import openpyxl

# def extract_file_references_and_aqs(directory, excel_path):
#     # Regex pattern to match file names with extensions (e.g., image1.jpg, script.js)
#     file_pattern = re.compile(r'[\w\-]+\.[a-zA-Z]{2,3}', re.IGNORECASE)
#     file_name_pattern = re.compile(r'[\w\-]+\.[a-zA-Z]{2,3}', re.IGNORECASE)
#     aqs_pattern = re.compile(r'<aqs:(\w+)', re.IGNORECASE)

#     # Dictionary to store unique file references grouped by their extensions
#     files_by_extension = {}
#     aqs_tags = set()

#     # Traverse the directory and process each file
#     for root, _, files in os.walk(directory):
#         for file in files:
#             file_path = os.path.join(root, file)

#             try:
#                 # Open and read the file content
#                 with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
#                     content = f.read()

#                 # Find all file references in the content
#                 references = file_name_pattern.findall(content)

#                 # Group the references by their extensions
#                 for ref in references:
#                     extension = os.path.splitext(ref)[1].lower()  # Extract extension
#                     if extension not in files_by_extension:
#                         files_by_extension[extension] = set()  # Use set to ensure uniqueness
#                     files_by_extension[extension].add(ref)

#                 # Extract aqs tags
#                 aqs_tags.update(aqs_pattern.findall(content))

#             except (UnicodeDecodeError, OSError):
#                 # Skip files that cannot be read as text
#                 continue

#     # Create an Excel workbook and add a worksheet
#     workbook = openpyxl.Workbook()
#     sheet = workbook.active
#     sheet.title = "Extracted Data"

#     # Write headers (file extensions as columns and AQS Tag column)
#     headers = list(files_by_extension.keys()) + ["AQS Tags"]
#     sheet.append(headers)

#     # Determine the maximum number of rows needed (files + aqs tags)
#     max_rows = max(len(files) for files in files_by_extension.values())  # Largest number of file names in a column
#     max_rows = max(max_rows, len(aqs_tags))  # Ensure there are enough rows for AQS tags

#     # Write unique file references and AQS tags into the Excel sheet
#     for i in range(max_rows):
#         row = []
        
#         # Add file names for each extension column
#         for extension in headers[:-1]:  # Exclude the "AQS Tags" column
#             file_list = list(files_by_extension.get(extension, []))
#             row.append(file_list[i] if i < len(file_list) else '')  # Add file name or empty cell
        
#         # Add AQS tags to the last column
#         row.append(list(aqs_tags)[i] if i < len(aqs_tags) else '')  # Add AQS tag or empty cell
#         sheet.append(row)

#     # Save the Excel workbook
#     workbook.save(excel_path)
#     print(f"Excel file saved at: {excel_path}")

# # Replace 'your_directory' with the path to the directory containing your files
# # Replace 'output.xlsx' with the desired path for the Excel file
# extract_file_references_and_aqs(r'C:\Users\ShashankTudum\Downloads\AQS_VBScript\AQS_VBScript', r'C:\Users\ShashankTudum\Downloads\shashank-output\output.xlsx')
"""
4th way
"""
# import os
# import re
# import openpyxl

# def extract_file_references(directory, excel_path):
#     # Generalized regex pattern to match any attribute with file paths and extensions
#     pattern = r'([a-zA-Z0-9-]+)="([^"]+\.[a-zA-Z]+)"'

#     # Dictionary to store unique file references grouped by their extensions
#     files_by_extension = {}

#     # Traverse the directory and process each file
#     for root, _, files in os.walk(directory):
#         for file in files:
#             file_path = os.path.join(root, file)

#             try:
#                 # Open and read the file content
#                 with open(file_path, 'r', encoding='utf-8') as f:
#                     content = f.read()

#                 # Find all file references in the content
#                 references = re.findall(pattern, content)

#                 # Group the references by their extensions (extracting filename with extension)
#                 for ref in references:
#                     file_name_with_extension = os.path.basename(ref[1])  # Extract filename with extension
#                     extension = os.path.splitext(file_name_with_extension)[1].lower()  # Extract extension
#                     if extension not in files_by_extension:
#                         files_by_extension[extension] = set()  # Use set to ensure uniqueness
#                     files_by_extension[extension].add(file_name_with_extension)

#             except (UnicodeDecodeError, OSError):
#                 # Skip files that cannot be read as text
#                 continue

#     # Create an Excel workbook and add a worksheet
#     workbook = openpyxl.Workbook()
#     sheet = workbook.active
#     sheet.title = "File References"

#     # Write headers (file extensions as columns)
#     headers = list(files_by_extension.keys())
#     sheet.append(headers)

#     # Determine the maximum number of rows needed (the highest number of files under any extension)
#     max_rows = max(len(files) for files in files_by_extension.values())

#     # Write the file names (under respective extension columns)
#     for i in range(max_rows):
#         row = []
#         for extension in headers:
#             file_list = list(files_by_extension[extension])
#             row.append(file_list[i] if i < len(file_list) else '')  # Add file name or leave empty
#         sheet.append(row)

#     # Save the Excel workbook
#     workbook.save(excel_path)
#     print(f"Excel file saved at: {excel_path}")

# # Replace 'your_directory' with the path to the directory containing your files
# # Replace 'output.xlsx' with the desired path for the Excel file
# extract_file_references(r'C:\Users\ShashankTudum\Downloads\AQS_VBScript\AQS_VBScript', r'C:\Users\ShashankTudum\Downloads\shashank-output\output.xlsx')
"""               """

import os
import re
import openpyxl

def extract_file_references(directory, excel_path):
    # Generalized regex pattern to match any attribute with file paths and extensions
    file_pattern = r'([a-zA-Z0-9-]+)="([^"]+\.[a-zA-Z]+)"'
    aqs_pattern = re.compile(r'<aqs:(\w+)', re.IGNORECASE)

    # Dictionary to store unique file references grouped by their extensions
    files_by_extension = {}
    aqs_tags = set()  # Set to store unique AQS tags

    # Traverse the directory and process each file
    for root, _, files in os.walk(directory):
        for file in files:
            file_path = os.path.join(root, file)

            try:
                # Open and read the file content
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()

                # Find all file references in the content
                references = re.findall(file_pattern, content)

                # Group the references by their extensions (extracting filename with extension)
                for ref in references:
                    file_name_with_extension = os.path.basename(ref[1])  # Extract filename with extension
                    extension = os.path.splitext(file_name_with_extension)[1].lower()  # Extract extension
                    if extension not in files_by_extension:
                        files_by_extension[extension] = set()  # Use set to ensure uniqueness
                    files_by_extension[extension].add(file_name_with_extension)

                # Find all AQS tags in the content
                aqs_matches = aqs_pattern.findall(content)
                aqs_tags.update(aqs_matches)

            except (UnicodeDecodeError, OSError):
                # Skip files that cannot be read as text
                continue

    # Create an Excel workbook and add a worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "File References"

    # Write headers (file extensions as columns) + AQS Tags
    headers = list(files_by_extension.keys())
    headers.append("AQS Tags")
    sheet.append(headers)

    # Determine the maximum number of rows needed (the highest number of files under any extension)
    max_rows = max((len(files) for files in files_by_extension.values()), default=0)
    max_rows = max(max_rows, len(aqs_tags))  # Include AQS tags count if larger

    # Write the file names (under respective extension columns) and AQS tags
    for i in range(max_rows):
        row = []
        for extension in headers[:-1]:  # Exclude "AQS Tags" column
            file_list = list(files_by_extension.get(extension, []))
            row.append(file_list[i] if i < len(file_list) else '')  # Add file name or leave empty
        aqs_tag_list = list(aqs_tags)
        row.append(aqs_tag_list[i] if i < len(aqs_tag_list) else '')  # Add AQS tag or leave empty
        sheet.append(row)

    # Save the Excel workbook
    workbook.save(excel_path)
    print(f"Excel file saved at: {excel_path}")

# Replace 'your_directory' with the path to the directory containing your files
# Replace 'output.xlsx' with the desired path for the Excel file
extract_file_references(
    r'C:\Users\ShashankTudum\Downloads\AQS_VBScript\AQS_VBScript',
    r'C:\Users\ShashankTudum\Downloads\shashank-output\output.xlsx'
)
